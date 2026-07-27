from __future__ import annotations

import argparse
import hashlib
import json
import os
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo


LOCAL_ZONE = ZoneInfo("America/Guayaquil")
COLLECTION_NAME = "boiler_consumption_readings"
SCHEMA_VERSION = 2
BOILER_MASTER_NAMES = {
    "alfa_laval_1200": "CalAlfa",
    "distral_900": "900Distral",
    "cleaver_brooks_1200": "CalCleaver",
}
SPANISH_WEEKDAYS = (
    "lunes",
    "martes",
    "miércoles",
    "jueves",
    "viernes",
    "sábado",
    "domingo",
)
SPANISH_MONTHS = (
    "",
    "ene",
    "feb",
    "mar",
    "abr",
    "may",
    "jun",
    "jul",
    "ago",
    "sep",
    "oct",
    "nov",
    "dic",
)


@dataclass(frozen=True)
class Reading:
    document_id: str
    boiler_id: str
    boiler_name: str
    recorded_at: datetime
    created_at: datetime | None
    operator_uid: str
    operator_name: str
    pressure_psi: float | None
    pressure_unit: str
    bunker_total_gal: float | None
    bunker_unit: str
    water_total_gal: float | None
    water_unit: str
    steam_total_gal: float | None
    steam_unit: str
    reading_mode: str
    revision: int
    root_record_id: str
    replaces_record_id: str
    status: str
    warnings: tuple[str, ...]
    notes: str
    app_version: str
    platform: str
    schema_version: int
    source: str


@dataclass(frozen=True)
class Interval:
    reading: Reading
    previous: Reading | None
    hours: float | None
    bunker_delta_gal: float | None
    water_delta_gal: float | None
    quality: str
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class HourlyAllocation:
    hour_start: datetime
    hour_end: datetime
    interval_start: datetime
    interval_end: datetime
    boiler_id: str
    source_document_id: str
    overlap_hours: float
    interval_hours: float
    bunker_interval_gal: float | None
    bunker_hour_gal: float | None
    water_interval_gal: float | None
    water_hour_gal: float | None
    pressure_average_psi: float | None
    quality: str


@dataclass(frozen=True)
class DailySummary:
    day: date
    boiler_id: str
    pressure_average_psi: float | None
    bunker_gal: float
    water_gal: float


def _as_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        parsed = value
    elif hasattr(value, "to_datetime"):
        parsed = value.to_datetime()
    else:
        text = str(value).strip().replace("Z", "+00:00")
        if not text:
            return None
        parsed = datetime.fromisoformat(text)
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _as_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def reading_from_dict(document_id: str, data: dict[str, Any]) -> Reading:
    recorded_at = _as_datetime(data.get("recordedAt"))
    if recorded_at is None:
        raise ValueError(f"{document_id}: recordedAt es obligatorio")
    boiler_id = str(data.get("boilerId") or "").strip()
    if not boiler_id:
        raise ValueError(f"{document_id}: boilerId es obligatorio")
    warnings = data.get("validationWarnings") or []
    if not isinstance(warnings, list):
        warnings = [str(warnings)]
    return Reading(
        document_id=document_id,
        boiler_id=boiler_id,
        boiler_name=str(
            data.get("boilerNameSnapshot") or data.get("boilerName") or boiler_id
        ),
        recorded_at=recorded_at,
        created_at=_as_datetime(data.get("createdAt")),
        operator_uid=str(data.get("createdByUid") or ""),
        operator_name=str(data.get("createdByNameSnapshot") or ""),
        pressure_psi=_as_float(data.get("boilerPressurePsi")),
        pressure_unit=str(data.get("boilerPressureUnit") or "psi"),
        bunker_total_gal=_as_float(data.get("bunkerValue", data.get("fuelTotal"))),
        bunker_unit=str(data.get("bunkerUnit") or ""),
        water_total_gal=_as_float(data.get("waterValue", data.get("waterTotal"))),
        water_unit=str(data.get("waterUnit") or ""),
        steam_total_gal=_as_float(data.get("steamValue", data.get("steamTotal"))),
        steam_unit=str(data.get("steamUnit") or ""),
        reading_mode=str(data.get("readingMode") or "cumulative_meter"),
        revision=_as_int(data.get("revision"), 1),
        root_record_id=str(data.get("rootRecordId") or document_id),
        replaces_record_id=str(data.get("replacesRecordId") or ""),
        status=str(data.get("status") or ""),
        warnings=tuple(str(item) for item in warnings),
        notes=str(data.get("notes") or ""),
        app_version=str(data.get("appVersion") or ""),
        platform=str(data.get("platform") or ""),
        schema_version=_as_int(data.get("schemaVersion"), 1),
        source=str(data.get("source") or "manual"),
    )


def latest_revisions(readings: Iterable[Reading]) -> list[Reading]:
    latest: dict[str, Reading] = {}
    for reading in readings:
        current = latest.get(reading.root_record_id)
        if current is None or (reading.revision, reading.created_at or reading.recorded_at) > (
            current.revision,
            current.created_at or current.recorded_at,
        ):
            latest[reading.root_record_id] = reading
    return sorted(latest.values(), key=lambda item: (item.boiler_id, item.recorded_at))


def _positive_delta(current: float | None, previous: float | None) -> float | None:
    if current is None or previous is None:
        return None
    delta = current - previous
    return delta if delta > 0 else None


def build_intervals(readings: Iterable[Reading]) -> list[Interval]:
    ordered = [
        reading
        for reading in latest_revisions(readings)
        if reading.reading_mode == "cumulative_meter"
    ]
    previous_by_boiler: dict[str, Reading] = {}
    result: list[Interval] = []
    for reading in ordered:
        previous = previous_by_boiler.get(reading.boiler_id)
        previous_by_boiler[reading.boiler_id] = reading
        if previous is None:
            result.append(
                Interval(
                    reading=reading,
                    previous=None,
                    hours=None,
                    bunker_delta_gal=None,
                    water_delta_gal=None,
                    quality="Primera lectura",
                    warnings=reading.warnings,
                )
            )
            continue
        hours = (reading.recorded_at - previous.recorded_at).total_seconds() / 3600
        bunker_units_valid = (
            reading.bunker_unit == "gal" and previous.bunker_unit == "gal"
        )
        water_units_valid = (
            reading.water_unit == "gal" and previous.water_unit == "gal"
        )
        bunker_delta = (
            _positive_delta(reading.bunker_total_gal, previous.bunker_total_gal)
            if bunker_units_valid
            else None
        )
        water_delta = (
            _positive_delta(reading.water_total_gal, previous.water_total_gal)
            if water_units_valid
            else None
        )
        warnings = list(reading.warnings)
        if not bunker_units_valid:
            warnings.append("bunker_unit_not_gal")
        if not water_units_valid:
            warnings.append("water_unit_not_gal")
        if hours <= 0:
            quality = "Fecha no válida"
            warnings.append("non_positive_interval")
        elif (
            bunker_units_valid
            and reading.bunker_total_gal is not None
            and previous.bunker_total_gal is not None
            and bunker_delta is None
        ) or (
            water_units_valid
            and reading.water_total_gal is not None
            and previous.water_total_gal is not None
            and water_delta is None
        ):
            quality = "Reinicio o lectura menor"
            warnings.append("meter_reset_or_non_positive_delta")
        elif abs(hours - 1) <= 0.1:
            quality = "Intervalo cercano a 1 h"
        else:
            quality = "Distribución proporcional"
        result.append(
            Interval(
                reading=reading,
                previous=previous,
                hours=hours,
                bunker_delta_gal=bunker_delta,
                water_delta_gal=water_delta,
                quality=quality,
                warnings=tuple(dict.fromkeys(warnings)),
            )
        )
    return result


def _average(values: Iterable[float | None]) -> float | None:
    available = [value for value in values if value is not None]
    return sum(available) / len(available) if available else None


def allocate_hourly(intervals: Iterable[Interval]) -> list[HourlyAllocation]:
    allocations: list[HourlyAllocation] = []
    for interval in intervals:
        if (
            interval.previous is None
            or interval.hours is None
            or interval.hours <= 0
            or (
                interval.bunker_delta_gal is None
                and interval.water_delta_gal is None
            )
        ):
            continue
        start = interval.previous.recorded_at.astimezone(LOCAL_ZONE)
        end = interval.reading.recorded_at.astimezone(LOCAL_ZONE)
        cursor = start.replace(minute=0, second=0, microsecond=0)
        pressure_average = _average(
            [interval.previous.pressure_psi, interval.reading.pressure_psi]
        )
        while cursor < end:
            hour_end = cursor + timedelta(hours=1)
            overlap_start = max(start, cursor)
            overlap_end = min(end, hour_end)
            overlap_hours = max(
                0.0, (overlap_end - overlap_start).total_seconds() / 3600
            )
            if overlap_hours > 0:
                share = overlap_hours / interval.hours
                allocations.append(
                    HourlyAllocation(
                        hour_start=cursor,
                        hour_end=hour_end,
                        interval_start=start,
                        interval_end=end,
                        boiler_id=interval.reading.boiler_id,
                        source_document_id=interval.reading.document_id,
                        overlap_hours=overlap_hours,
                        interval_hours=interval.hours,
                        bunker_interval_gal=interval.bunker_delta_gal,
                        bunker_hour_gal=(
                            interval.bunker_delta_gal * share
                            if interval.bunker_delta_gal is not None
                            else None
                        ),
                        water_interval_gal=interval.water_delta_gal,
                        water_hour_gal=(
                            interval.water_delta_gal * share
                            if interval.water_delta_gal is not None
                            else None
                        ),
                        pressure_average_psi=pressure_average,
                        quality=interval.quality,
                    )
                )
            cursor = hour_end
    return allocations


def build_daily_summaries(
    readings: Iterable[Reading], allocations: Iterable[HourlyAllocation]
) -> list[DailySummary]:
    latest = [
        reading
        for reading in latest_revisions(readings)
        if reading.reading_mode == "cumulative_meter"
    ]
    pressure_values: dict[tuple[date, str], list[float]] = defaultdict(list)
    keys: set[tuple[date, str]] = set()
    for reading in latest:
        local_day = reading.recorded_at.astimezone(LOCAL_ZONE).date()
        key = (local_day, reading.boiler_id)
        keys.add(key)
        if reading.pressure_psi is not None:
            pressure_values[key].append(reading.pressure_psi)
    bunker_totals: dict[tuple[date, str], float] = defaultdict(float)
    water_totals: dict[tuple[date, str], float] = defaultdict(float)
    for allocation in allocations:
        key = (allocation.hour_start.date(), allocation.boiler_id)
        keys.add(key)
        if allocation.bunker_hour_gal is not None:
            bunker_totals[key] += allocation.bunker_hour_gal
        if allocation.water_hour_gal is not None:
            water_totals[key] += allocation.water_hour_gal
    return [
        DailySummary(
            day=day,
            boiler_id=boiler_id,
            pressure_average_psi=_average(pressure_values[(day, boiler_id)]),
            bunker_gal=bunker_totals[(day, boiler_id)],
            water_gal=water_totals[(day, boiler_id)],
        )
        for day, boiler_id in sorted(keys)
    ]


def _local_date_bounds_utc(day: date) -> tuple[datetime, datetime]:
    start = datetime(
        day.year,
        day.month,
        day.day,
        tzinfo=LOCAL_ZONE,
    ).astimezone(timezone.utc)
    return start, start + timedelta(days=1)


def load_firestore_readings(
    *,
    changed_since_utc: datetime | None = None,
    target_local_date: date | None = None,
) -> list[Reading]:
    import firebase_admin
    from firebase_admin import credentials, firestore
    from google.cloud.firestore_v1.base_query import FieldFilter

    if not firebase_admin._apps:
        credential_json = os.getenv("FIREBASE_SERVICE_ACCOUNT_JSON")
        if credential_json:
            firebase_admin.initialize_app(
                credentials.Certificate(json.loads(credential_json))
            )
        else:
            firebase_admin.initialize_app()
    client = firestore.client()
    collection = client.collection(COLLECTION_NAME)
    readings_by_id: dict[str, Reading] = {}

    def add_documents(documents: Iterable[Any]) -> None:
        for document in documents:
            try:
                readings_by_id[document.id] = reading_from_dict(
                    document.id,
                    document.to_dict(),
                )
            except ValueError as error:
                print(f"ADVERTENCIA: {error}")

    if target_local_date is None and changed_since_utc is None:
        add_documents(collection.order_by("recordedAt").stream())
        return sorted(
            readings_by_id.values(),
            key=lambda reading: (reading.recorded_at, reading.document_id),
        )

    affected_dates: set[date]
    affected_boilers: set[str]
    if target_local_date is not None:
        start_utc, end_utc = _local_date_bounds_utc(target_local_date)
        day_documents = (
            collection.where(
                filter=FieldFilter("recordedAt", ">=", start_utc)
            )
            .where(filter=FieldFilter("recordedAt", "<", end_utc))
            .order_by("recordedAt")
            .stream()
        )
        add_documents(day_documents)
        if not readings_by_id:
            return []
        affected_dates = {target_local_date}
        affected_boilers = {
            reading.boiler_id for reading in readings_by_id.values()
        }
    else:
        cutoff = changed_since_utc.astimezone(timezone.utc)
        changed_documents = (
            collection.where(filter=FieldFilter("createdAt", ">", cutoff))
            .order_by("createdAt")
            .stream()
        )
        add_documents(changed_documents)
        if not readings_by_id:
            return []
        affected_dates = {
            reading.recorded_at.astimezone(LOCAL_ZONE).date()
            for reading in readings_by_id.values()
        }
        affected_boilers = {
            reading.boiler_id for reading in readings_by_id.values()
        }

    context_start, _ = _local_date_bounds_utc(min(affected_dates))
    _, context_end = _local_date_bounds_utc(max(affected_dates))
    context_documents = (
        collection.where(
            filter=FieldFilter("recordedAt", ">=", context_start)
        )
        .where(filter=FieldFilter("recordedAt", "<", context_end))
        .order_by("recordedAt")
        .stream()
    )
    add_documents(context_documents)

    for boiler_id in sorted(affected_boilers):
        previous_documents = (
            collection.where(filter=FieldFilter("boilerId", "==", boiler_id))
            .where(filter=FieldFilter("recordedAt", "<", context_start))
            .order_by("recordedAt", direction=firestore.Query.DESCENDING)
            .limit(1)
            .stream()
        )
        add_documents(previous_documents)
        next_documents = (
            collection.where(filter=FieldFilter("boilerId", "==", boiler_id))
            .where(filter=FieldFilter("recordedAt", ">=", context_end))
            .order_by("recordedAt")
            .limit(1)
            .stream()
        )
        add_documents(next_documents)

    return sorted(
        readings_by_id.values(),
        key=lambda reading: (reading.recorded_at, reading.document_id),
    )


def load_json_readings(input_path: Path) -> list[Reading]:
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError("El JSON de entrada debe contener una lista de documentos")
    readings: list[Reading] = []
    for index, item in enumerate(payload):
        document_id = str(item.get("documentId") or item.get("id") or f"row-{index}")
        readings.append(reading_from_dict(document_id, item))
    return readings


def _excel_datetime(value: datetime | None, local: bool = False) -> datetime | None:
    if value is None:
        return None
    converted = value.astimezone(LOCAL_ZONE if local else timezone.utc)
    return converted.replace(tzinfo=None)


def excel_weeknum_sunday(day: date) -> int:
    first_day = date(day.year, 1, 1)
    sunday_offset = (first_day.weekday() + 1) % 7
    return ((day - first_day).days + sunday_offset) // 7 + 1


def build_workbook(readings: list[Reading], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    intervals = build_intervals(readings)
    allocations = allocate_hourly(intervals)
    summaries = build_daily_summaries(readings, allocations)
    workbook = Workbook()
    workbook.remove(workbook.active)
    control = workbook.create_sheet("Control")
    raw = workbook.create_sheet("Datos_Firestore")
    interval_sheet = workbook.create_sheet("Intervalos")
    hourly_sheet = workbook.create_sheet("Consumo_Horario")
    daily_sheet = workbook.create_sheet("Resumen_Diario")

    black = "1F1F1F"
    white = "FFFFFF"
    blue = "D9EAF7"
    light_blue = "EDF5FA"
    dark_blue = "1F4E78"
    yellow = "FFF2CC"
    border_color = "B7C9D6"
    thin = Side(style="thin", color=border_color)

    def style_header(sheet, row: int, last_column: int) -> None:
        for cell in sheet[row][:last_column]:
            cell.fill = PatternFill("solid", fgColor=black)
            cell.font = Font(bold=True, color=white)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_rows(sheet, start_row: int, end_row: int, last_column: int) -> None:
        for row in sheet.iter_rows(
            min_row=start_row, max_row=end_row, min_col=1, max_col=last_column
        ):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=light_blue)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center")

    def set_widths(sheet, widths: list[int]) -> None:
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    control.merge_cells("A1:F1")
    control["A1"] = "Control de integración - Consumos de calderas"
    control["A1"].fill = PatternFill("solid", fgColor=dark_blue)
    control["A1"].font = Font(bold=True, color=white, size=16)
    control["A1"].alignment = Alignment(vertical="center")
    control.row_dimensions[1].height = 30
    metadata = [
        ("Estado", "Sincronización completada"),
        ("Zona horaria", "America/Guayaquil"),
        ("Origen", "Firebase Cloud Firestore"),
        ("Colección", COLLECTION_NAME),
        ("Registros crudos", len(readings)),
        ("Lecturas efectivas", len(latest_revisions(readings))),
        ("Intervalos", len(intervals)),
        ("Asignaciones horarias", len(allocations)),
        ("Filas diarias", len(summaries)),
        ("Última sincronización UTC", datetime.now(timezone.utc).isoformat()),
        ("Versión del exportador", SCHEMA_VERSION),
    ]
    for row, (label, value) in enumerate(metadata, 3):
        control.cell(row=row, column=1, value=label)
        control.cell(row=row, column=2, value=value)
    style_header(control, 3, 1)
    for row in range(4, 3 + len(metadata)):
        control.cell(row=row, column=1).fill = PatternFill("solid", fgColor=black)
        control.cell(row=row, column=1).font = Font(bold=True, color=white)
    style_rows(control, 3, 2 + len(metadata), 2)
    for row in range(3, 3 + len(metadata)):
        control.cell(row=row, column=1).fill = PatternFill("solid", fgColor=black)
        control.cell(row=row, column=1).font = Font(bold=True, color=white)
    set_widths(control, [30, 78, 12, 12, 12, 12])

    raw_headers = [
        "documentId",
        "boilerId",
        "boilerName",
        "recordedAtUtc",
        "recordedAtLocal",
        "createdAtUtc",
        "operatorUid",
        "operatorName",
        "boilerPressurePsi",
        "pressureUnit",
        "bunkerCumulativeGal",
        "bunkerUnit",
        "waterCumulativeGal",
        "waterUnit",
        "steamCumulativeGal",
        "steamUnit",
        "readingMode",
        "revision",
        "rootRecordId",
        "replacesRecordId",
        "status",
        "warnings",
        "notes",
        "appVersion",
        "platform",
        "schemaVersion",
        "source",
    ]
    raw.append(raw_headers)
    for reading in sorted(readings, key=lambda item: (item.boiler_id, item.recorded_at)):
        raw.append(
            [
                reading.document_id,
                reading.boiler_id,
                reading.boiler_name,
                _excel_datetime(reading.recorded_at),
                _excel_datetime(reading.recorded_at, local=True),
                _excel_datetime(reading.created_at),
                reading.operator_uid,
                reading.operator_name,
                reading.pressure_psi,
                reading.pressure_unit,
                reading.bunker_total_gal,
                reading.bunker_unit,
                reading.water_total_gal,
                reading.water_unit,
                reading.steam_total_gal,
                reading.steam_unit,
                reading.reading_mode,
                reading.revision,
                reading.root_record_id,
                reading.replaces_record_id,
                reading.status,
                " | ".join(reading.warnings),
                reading.notes,
                reading.app_version,
                reading.platform,
                reading.schema_version,
                reading.source,
            ]
        )
    style_header(raw, 1, len(raw_headers))
    if raw.max_row > 1:
        style_rows(raw, 2, raw.max_row, len(raw_headers))
    raw.freeze_panes = "A2"
    raw.auto_filter.ref = raw.dimensions
    for row in range(2, raw.max_row + 1):
        for column in range(4, 7):
            raw.cell(row=row, column=column).number_format = "yyyy-mm-dd hh:mm:ss"
    set_widths(
        raw,
        [27, 24, 28, 21, 21, 21, 25, 28, 18, 14, 22, 14, 22, 14, 22, 14, 20, 10, 27, 27, 15, 34, 36, 14, 12, 14, 14],
    )

    interval_headers = [
        "documentId",
        "boilerId",
        "Caldera master",
        "Fecha y hora lectura",
        "Fecha y hora anterior",
        "Horas transcurridas",
        "Búnker acumulado (gal)",
        "Búnker anterior (gal)",
        "Delta búnker válido (gal)",
        "Agua acumulada (gal)",
        "Agua anterior (gal)",
        "Delta agua válido (gal)",
        "Presión (PSI)",
        "Calidad",
        "Advertencias",
    ]
    interval_sheet.append(interval_headers)
    for interval in intervals:
        previous = interval.previous
        interval_sheet.append(
            [
                interval.reading.document_id,
                interval.reading.boiler_id,
                BOILER_MASTER_NAMES.get(interval.reading.boiler_id, "REVISAR"),
                _excel_datetime(interval.reading.recorded_at, local=True),
                _excel_datetime(previous.recorded_at, local=True) if previous else None,
                interval.hours,
                interval.reading.bunker_total_gal,
                previous.bunker_total_gal if previous else None,
                interval.bunker_delta_gal,
                interval.reading.water_total_gal,
                previous.water_total_gal if previous else None,
                interval.water_delta_gal,
                interval.reading.pressure_psi,
                interval.quality,
                " | ".join(interval.warnings),
            ]
        )
    style_header(interval_sheet, 1, len(interval_headers))
    if interval_sheet.max_row > 1:
        style_rows(interval_sheet, 2, interval_sheet.max_row, len(interval_headers))
    interval_sheet.freeze_panes = "A2"
    interval_sheet.auto_filter.ref = interval_sheet.dimensions
    set_widths(
        interval_sheet,
        [27, 24, 20, 22, 22, 18, 22, 21, 22, 22, 21, 22, 15, 29, 38],
    )

    hourly_headers = [
        "Hora inicio",
        "Hora fin",
        "Fecha",
        "boilerId",
        "Caldera master",
        "Intervalo inicio",
        "Intervalo fin",
        "Horas solapadas",
        "Horas intervalo",
        "Delta búnker intervalo",
        "Consumo búnker hora (gal)",
        "Delta agua intervalo",
        "Consumo agua hora (gal)",
        "Presión promedio (PSI)",
        "Calidad",
        "sourceDocumentId",
    ]
    hourly_sheet.append(hourly_headers)
    for item in allocations:
        hourly_sheet.append(
            [
                _excel_datetime(item.hour_start, local=True),
                _excel_datetime(item.hour_end, local=True),
                item.hour_start.date(),
                item.boiler_id,
                BOILER_MASTER_NAMES.get(item.boiler_id, "REVISAR"),
                _excel_datetime(item.interval_start, local=True),
                _excel_datetime(item.interval_end, local=True),
                item.overlap_hours,
                item.interval_hours,
                item.bunker_interval_gal,
                item.bunker_hour_gal,
                item.water_interval_gal,
                item.water_hour_gal,
                item.pressure_average_psi,
                item.quality,
                item.source_document_id,
            ]
        )
    style_header(hourly_sheet, 1, len(hourly_headers))
    if hourly_sheet.max_row > 1:
        style_rows(hourly_sheet, 2, hourly_sheet.max_row, len(hourly_headers))
    hourly_sheet.freeze_panes = "A2"
    hourly_sheet.auto_filter.ref = hourly_sheet.dimensions
    set_widths(
        hourly_sheet,
        [20, 20, 13, 24, 20, 20, 20, 18, 17, 22, 24, 21, 23, 21, 28, 28],
    )

    daily_headers = [
        "Fecha",
        "Día",
        "Año",
        "Mes",
        "Presión",
        "Semana",
        "Caldera",
        "Minutos de Paro Total",
        "Horas Funcionamiento OPTIMO",
        "Tiempo FUNCIONAMIENTO REAL",
        "Eficiencia",
        "Alcance",
        "CONSUMO DE BUNKER(glns)",
        "CONSUMO DE AGUA (glns)",
        "Ton/Agua",
        "galones/m3",
    ]
    daily_sheet.append(daily_headers)
    for summary in summaries:
        water_tons = summary.water_gal * 3.785 / 1000
        daily_sheet.append(
            [
                summary.day,
                SPANISH_WEEKDAYS[summary.day.weekday()],
                summary.day.year,
                SPANISH_MONTHS[summary.day.month],
                summary.pressure_average_psi,
                f"S {excel_weeknum_sunday(summary.day)}",
                BOILER_MASTER_NAMES.get(summary.boiler_id, "REVISAR"),
                0,
                24,
                24,
                1,
                0.98,
                summary.bunker_gal,
                summary.water_gal,
                water_tons,
                summary.bunker_gal / water_tons if water_tons > 0 else None,
            ]
        )
    style_header(daily_sheet, 1, len(daily_headers))
    if daily_sheet.max_row > 1:
        style_rows(daily_sheet, 2, daily_sheet.max_row, len(daily_headers))
        for row in range(2, daily_sheet.max_row + 1):
            for column in range(9, 13):
                daily_sheet.cell(row=row, column=column).fill = PatternFill(
                    "solid", fgColor=yellow
                )
    daily_sheet.freeze_panes = "A2"
    daily_sheet.auto_filter.ref = daily_sheet.dimensions
    set_widths(
        daily_sheet,
        [15, 14, 10, 11, 13, 13, 20, 22, 24, 24, 13, 13, 24, 24, 14, 16],
    )
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_power_automate_payload(
    readings: list[Reading],
    *,
    changed_since_utc: datetime | None = None,
    target_local_date: date | None = None,
    cursor_start_utc: datetime | None = None,
    cursor_end_utc: datetime | None = None,
    remaining_dates: Iterable[date] = (),
) -> dict[str, Any]:
    intervals = build_intervals(readings)
    allocations = allocate_hourly(intervals)
    summaries = build_daily_summaries(readings, allocations)
    payload_readings = readings
    affected_days: set[date] | None = None
    if target_local_date is not None:
        affected_days = {target_local_date}
        payload_readings = [
            reading
            for reading in readings
            if reading.recorded_at.astimezone(LOCAL_ZONE).date()
            == target_local_date
        ]
        intervals = [
            interval
            for interval in intervals
            if interval.reading.recorded_at.astimezone(LOCAL_ZONE).date()
            == target_local_date
        ]
        allocations = [
            allocation
            for allocation in allocations
            if allocation.hour_start.astimezone(LOCAL_ZONE).date()
            == target_local_date
        ]
        summaries = [
            summary for summary in summaries if summary.day == target_local_date
        ]
    elif changed_since_utc is not None:
        cutoff = changed_since_utc.astimezone(timezone.utc)
        payload_readings = [
            reading
            for reading in readings
            if reading.recorded_at >= cutoff
            or (reading.created_at is not None and reading.created_at >= cutoff)
        ]
        affected_days = {
            reading.recorded_at.astimezone(LOCAL_ZONE).date()
            for reading in payload_readings
        }
        intervals = [
            interval
            for interval in intervals
            if interval.reading.recorded_at.astimezone(LOCAL_ZONE).date()
            in affected_days
        ]
        allocations = [
            allocation
            for allocation in allocations
            if allocation.hour_start.astimezone(LOCAL_ZONE).date()
            in affected_days
        ]
        summaries = [
            summary for summary in summaries if summary.day in affected_days
        ]
    raw_headers = [
        "documentId",
        "boilerId",
        "boilerName",
        "recordedAtUtc",
        "recordedAtLocal",
        "createdAtUtc",
        "operatorUid",
        "operatorName",
        "boilerPressurePsi",
        "pressureUnit",
        "bunkerCumulativeGal",
        "bunkerUnit",
        "waterCumulativeGal",
        "waterUnit",
        "steamCumulativeGal",
        "steamUnit",
        "readingMode",
        "revision",
        "rootRecordId",
        "replacesRecordId",
        "status",
        "warnings",
        "notes",
        "appVersion",
        "platform",
        "schemaVersion",
        "source",
    ]
    raw_rows = [
        [
            reading.document_id,
            reading.boiler_id,
            reading.boiler_name,
            reading.recorded_at.isoformat(),
            reading.recorded_at.astimezone(LOCAL_ZONE).isoformat(),
            reading.created_at.isoformat() if reading.created_at else "",
            reading.operator_uid,
            reading.operator_name,
            reading.pressure_psi,
            reading.pressure_unit,
            reading.bunker_total_gal,
            reading.bunker_unit,
            reading.water_total_gal,
            reading.water_unit,
            reading.steam_total_gal,
            reading.steam_unit,
            reading.reading_mode,
            reading.revision,
            reading.root_record_id,
            reading.replaces_record_id,
            reading.status,
            " | ".join(reading.warnings),
            reading.notes,
            reading.app_version,
            reading.platform,
            reading.schema_version,
            reading.source,
        ]
        for reading in sorted(
            payload_readings, key=lambda item: (item.boiler_id, item.recorded_at)
        )
    ]
    interval_headers = [
        "documentId",
        "boilerId",
        "Caldera master",
        "Fecha y hora lectura",
        "Fecha y hora anterior",
        "Horas transcurridas",
        "Búnker acumulado (gal)",
        "Búnker anterior (gal)",
        "Delta búnker válido (gal)",
        "Agua acumulada (gal)",
        "Agua anterior (gal)",
        "Delta agua válido (gal)",
        "Presión (PSI)",
        "Calidad",
        "Advertencias",
    ]
    interval_rows = [
        [
            interval.reading.document_id,
            interval.reading.boiler_id,
            BOILER_MASTER_NAMES.get(interval.reading.boiler_id, "REVISAR"),
            interval.reading.recorded_at.astimezone(LOCAL_ZONE).isoformat(),
            (
                interval.previous.recorded_at.astimezone(LOCAL_ZONE).isoformat()
                if interval.previous
                else ""
            ),
            interval.hours,
            interval.reading.bunker_total_gal,
            interval.previous.bunker_total_gal if interval.previous else None,
            interval.bunker_delta_gal,
            interval.reading.water_total_gal,
            interval.previous.water_total_gal if interval.previous else None,
            interval.water_delta_gal,
            interval.reading.pressure_psi,
            interval.quality,
            " | ".join(interval.warnings),
        ]
        for interval in intervals
    ]
    hourly_headers = [
        "Hora inicio",
        "Hora fin",
        "Fecha",
        "boilerId",
        "Caldera master",
        "Intervalo inicio",
        "Intervalo fin",
        "Horas solapadas",
        "Horas intervalo",
        "Delta búnker intervalo",
        "Consumo búnker hora (gal)",
        "Delta agua intervalo",
        "Consumo agua hora (gal)",
        "Presión promedio (PSI)",
        "Calidad",
        "sourceDocumentId",
    ]
    hourly_rows = [
        [
            item.hour_start.isoformat(),
            item.hour_end.isoformat(),
            item.hour_start.date().isoformat(),
            item.boiler_id,
            BOILER_MASTER_NAMES.get(item.boiler_id, "REVISAR"),
            item.interval_start.isoformat(),
            item.interval_end.isoformat(),
            item.overlap_hours,
            item.interval_hours,
            item.bunker_interval_gal,
            item.bunker_hour_gal,
            item.water_interval_gal,
            item.water_hour_gal,
            item.pressure_average_psi,
            item.quality,
            item.source_document_id,
        ]
        for item in allocations
    ]
    daily_headers = [
        "syncKey",
        "Fecha",
        "Día",
        "Año",
        "Mes",
        "Presión",
        "Semana",
        "Caldera",
        "Minutos de Paro Total",
        "Horas Funcionamiento OPTIMO",
        "Tiempo FUNCIONAMIENTO REAL",
        "Eficiencia",
        "Alcance",
        "CONSUMO DE BUNKER(glns)",
        "CONSUMO DE AGUA (glns)",
        "Ton/Agua",
        "galones/m3",
    ]
    daily_rows: list[list[Any]] = []
    for summary in summaries:
        water_tons = summary.water_gal * 3.785 / 1000
        boiler_name = BOILER_MASTER_NAMES.get(summary.boiler_id, "REVISAR")
        daily_rows.append(
            [
                f"{summary.day.isoformat()}|{summary.boiler_id}",
                summary.day.isoformat(),
                SPANISH_WEEKDAYS[summary.day.weekday()],
                summary.day.year,
                SPANISH_MONTHS[summary.day.month],
                summary.pressure_average_psi,
                f"S {excel_weeknum_sunday(summary.day)}",
                boiler_name,
                0,
                24,
                24,
                1,
                0.98,
                summary.bunker_gal,
                summary.water_gal,
                water_tons,
                summary.bunker_gal / water_tons if water_tons > 0 else None,
            ]
        )
    generated_at_utc = datetime.now(timezone.utc).isoformat()
    cursor_candidates = [
        reading.created_at or reading.recorded_at for reading in payload_readings
    ]
    resolved_cursor_end_utc = cursor_end_utc or (
        max(cursor_candidates).astimezone(timezone.utc)
        if cursor_candidates
        else cursor_start_utc
    )
    cursor_start_text = (
        cursor_start_utc.astimezone(timezone.utc).isoformat()
        if cursor_start_utc is not None
        else ""
    )
    cursor_end_text = (
        resolved_cursor_end_utc.astimezone(timezone.utc).isoformat()
        if resolved_cursor_end_utc
        else ""
    )
    remaining_date_values = sorted(set(remaining_dates))
    batch_seed = "|".join(
        [
            target_local_date.isoformat() if target_local_date else "incremental",
            cursor_start_text,
            cursor_end_text,
            *sorted(reading.document_id for reading in payload_readings),
        ]
    )
    batch_id = hashlib.sha256(batch_seed.encode("utf-8")).hexdigest()[:20]
    return {
        "schemaVersion": SCHEMA_VERSION,
        "generatedAtUtc": generated_at_utc,
        "timezone": "America/Guayaquil",
        "collection": COLLECTION_NAME,
        "mode": (
            "date"
            if target_local_date is not None
            else "incremental"
            if changed_since_utc
            else "full"
        ),
        "windowStartUtc": (
            _local_date_bounds_utc(target_local_date)[0].isoformat()
            if target_local_date is not None
            else changed_since_utc.astimezone(timezone.utc).isoformat()
            if changed_since_utc is not None
            else ""
        ),
        "affectedDates": (
            sorted(day.isoformat() for day in affected_days)
            if affected_days is not None
            else []
        ),
        "delivery": {
            "status": "ready",
            "batchId": batch_id,
            "cursorStartUtc": cursor_start_text,
            "cursorEndUtc": cursor_end_text,
            "targetLocalDate": (
                target_local_date.isoformat() if target_local_date else ""
            ),
            "rawRowCount": len(raw_rows),
            "remainingDates": [
                pending_date.isoformat()
                for pending_date in remaining_date_values
            ],
        },
        "datasets": {
            "raw": {"headers": raw_headers, "rows": raw_rows},
            "intervals": {"headers": interval_headers, "rows": interval_rows},
            "hourly": {"headers": hourly_headers, "rows": hourly_rows},
            "daily": {"headers": daily_headers, "rows": daily_rows},
        },
    }


def start_of_local_day_utc(now_utc: datetime | None = None) -> datetime:
    current = (now_utc or datetime.now(timezone.utc)).astimezone(LOCAL_ZONE)
    return current.replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
    ).astimezone(timezone.utc)


def load_github_issue_payload() -> dict[str, Any] | None:
    import requests

    repository = os.environ["EE_DATA_REPOSITORY"].strip()
    issue_number = int(os.environ.get("EE_DATA_ISSUE_NUMBER", "1"))
    token = os.environ["EE_DATA_REPO_TOKEN"].strip()
    response = requests.get(
        f"https://api.github.com/repos/{repository}/issues/{issue_number}",
        headers={
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28",
        },
        timeout=180,
    )
    response.raise_for_status()
    body = response.json().get("body")
    if not isinstance(body, str) or not body.strip():
        return None
    try:
        payload = json.loads(body)
    except json.JSONDecodeError:
        return None
    return payload if isinstance(payload, dict) else None


def delivery_status(payload: dict[str, Any] | None) -> str:
    if not payload:
        return ""
    delivery = payload.get("delivery")
    if not isinstance(delivery, dict):
        return ""
    return str(delivery.get("status") or "").strip().lower()


def delivery_remaining_dates(
    payload: dict[str, Any] | None,
) -> list[date]:
    if not payload:
        return []
    delivery = payload.get("delivery")
    if not isinstance(delivery, dict):
        return []
    values = delivery.get("remainingDates")
    if not isinstance(values, list):
        return []
    result: list[date] = []
    for value in values:
        if isinstance(value, str):
            try:
                result.append(date.fromisoformat(value))
            except ValueError:
                continue
    return sorted(set(result))


def delivery_cursor_end_utc(
    payload: dict[str, Any] | None,
) -> datetime | None:
    if not payload:
        return None
    delivery = payload.get("delivery")
    if not isinstance(delivery, dict):
        return None
    raw_cursor = delivery.get("cursorEndUtc")
    if not isinstance(raw_cursor, str) or not raw_cursor.strip():
        return None
    parsed = datetime.fromisoformat(raw_cursor.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def acknowledged_cursor_utc(
    payload: dict[str, Any] | None,
) -> datetime | None:
    if delivery_status(payload) != "acknowledged":
        return None
    delivery = payload.get("delivery")
    if not isinstance(delivery, dict):
        return None
    raw_cursor = delivery.get(
        "cursorStartUtc"
        if delivery_remaining_dates(payload)
        else "cursorEndUtc"
    )
    if not isinstance(raw_cursor, str) or not raw_cursor.strip():
        return None
    parsed = datetime.fromisoformat(raw_cursor.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def payload_size_chars(payload: dict[str, Any]) -> int:
    return len(
        json.dumps(
            payload,
            ensure_ascii=False,
            separators=(",", ":"),
        )
    )


def publish_payload_to_github_issue(payload: dict[str, Any]) -> None:
    import requests

    repository = os.environ["EE_DATA_REPOSITORY"].strip()
    issue_number = int(os.environ.get("EE_DATA_ISSUE_NUMBER", "1"))
    token = os.environ["EE_DATA_REPO_TOKEN"].strip()
    body = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
    )
    max_chars = int(os.environ.get("EE_DATA_ISSUE_MAX_CHARS", "60000"))
    if len(body) > max_chars:
        raise ValueError(
            "El payload supera el limite seguro del issue de GitHub: "
            f"{len(body)} > {max_chars} caracteres"
        )
    response = requests.patch(
        f"https://api.github.com/repos/{repository}/issues/{issue_number}",
        headers={
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "X-GitHub-Api-Version": "2022-11-28",
        },
        json={"body": body},
        timeout=180,
    )
    response.raise_for_status()


def send_to_power_automate(payload: dict[str, Any]) -> None:
    """Backward-compatible alias for the GitHub issue transport."""
    publish_payload_to_github_issue(payload)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Exporta lecturas acumuladas de calderas desde Firestore a Excel."
    )
    parser.add_argument("--input-json", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument(
        "--skip-power-automate",
        "--skip-publish",
        dest="skip_publish",
        action="store_true",
    )
    parser.add_argument("--full-payload", action="store_true")
    parser.add_argument(
        "--payload-date",
        type=date.fromisoformat,
        help="Publica solamente la fecha local indicada (AAAA-MM-DD).",
    )
    args = parser.parse_args()
    if args.full_payload and args.payload_date is not None:
        parser.error("--full-payload y --payload-date no pueden combinarse")

    issue_payload: dict[str, Any] | None = None
    if not args.skip_publish and args.input_json is None:
        issue_payload = load_github_issue_payload()
        if delivery_status(issue_payload) == "ready":
            delivery = issue_payload.get("delivery", {})
            print(
                "Hay un lote pendiente de confirmacion en Power Automate; "
                f"no se consulto Firestore ni se sobrescribio el issue "
                f"(batch {delivery.get('batchId', 'sin-id')})."
            )
            return

    acknowledged_cursor = acknowledged_cursor_utc(issue_payload)
    queued_dates = delivery_remaining_dates(issue_payload)
    target_local_date = args.payload_date or (
        queued_dates[0] if queued_dates else None
    )
    remaining_dates = (
        queued_dates[1:]
        if args.payload_date is None and queued_dates
        else []
    )
    queued_cursor_end = (
        delivery_cursor_end_utc(issue_payload) if queued_dates else None
    )
    changed_since_utc = (
        None
        if args.full_payload or target_local_date is not None
        else acknowledged_cursor or start_of_local_day_utc()
    )
    readings = (
        load_json_readings(args.input_json)
        if args.input_json
        else load_firestore_readings(
            changed_since_utc=changed_since_utc,
            target_local_date=target_local_date,
        )
    )
    if args.output:
        build_workbook(readings, args.output)
        print(f"Libro auxiliar de diagnóstico generado: {args.output}")
    if not args.skip_publish:
        payload = build_power_automate_payload(
            readings,
            changed_since_utc=changed_since_utc,
            target_local_date=target_local_date,
            cursor_start_utc=acknowledged_cursor,
            cursor_end_utc=queued_cursor_end,
            remaining_dates=remaining_dates,
        )
        if not payload["datasets"]["raw"]["rows"]:
            print(
                "No hay lecturas nuevas para publicar; "
                "el issue privado se mantuvo sin cambios."
            )
            return
        max_chars = int(os.environ.get("EE_DATA_ISSUE_MAX_CHARS", "60000"))
        if payload_size_chars(payload) > max_chars and target_local_date is None:
            affected_dates = [
                date.fromisoformat(value)
                for value in payload.get("affectedDates", [])
            ]
            if len(affected_dates) > 1:
                first_date = affected_dates[0]
                payload = build_power_automate_payload(
                    readings,
                    target_local_date=first_date,
                    cursor_start_utc=acknowledged_cursor,
                    cursor_end_utc=delivery_cursor_end_utc(payload),
                    remaining_dates=affected_dates[1:],
                )
                print(
                    "El lote incremental se dividio automaticamente por fecha; "
                    f"se publica {first_date.isoformat()} y quedan "
                    f"{len(affected_dates) - 1} fecha(s) en cola."
                )
        publish_payload_to_github_issue(payload)
        print(
            "Payload publicado en el issue privado de GitHub "
            f"({payload_size_chars(payload)} caracteres)"
        )


if __name__ == "__main__":
    main()
